from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Sum, Count, Avg, F, FloatField
from django.db.models.functions import TruncWeek, TruncMonth, ExtractWeek, ExtractYear
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
import json
import calendar
from portal.models import (
    DailyReportingModel, ActivityReportingModel, Activities, staffTbl,
    PersonnelModel, FarmdetailsTbl, Community, cocoaDistrict, projectTbl
)

@login_required
def weekly_monitoring(request):
    """Render the Weekly Monitoring dashboard"""
    context = {
        'weeks': get_last_12_weeks(),
        'districts': cocoaDistrict.objects.filter(delete_field='no').values('id', 'name'),
        'activities': Activities.objects.values('id', 'main_activity').distinct(),
    }
    return render(request, 'portal/activity_reporting/weekly_monitoring.html', context)

def get_last_12_weeks():
    """Get the last 12 weeks for dropdown"""
    weeks = []
    today = timezone.now().date()
    for i in range(12):
        week_start = today - timedelta(days=today.weekday() + 7 * i)
        week_end = week_start + timedelta(days=6)
        week_num = week_start.isocalendar()[1]
        year = week_start.year
        weeks.append({
            'value': f"{year}-W{week_num}",
            'label': f"Week {week_num}, {year} ({week_start.strftime('%d %b')} - {week_end.strftime('%d %b')})",
            'start': week_start.strftime('%Y-%m-%d'),
            'end': week_end.strftime('%Y-%m-%d')
        })
    return weeks

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_summary(request):
    """Get weekly summary statistics"""
    try:
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        district_id = request.GET.get('district_id')
        po_id = request.GET.get('po_id')
        
        # Base queryset for daily reports
        queryset = DailyReportingModel.objects.filter(delete_field='no')
        
        # Apply date filter
        if week_start and week_end:
            queryset = queryset.filter(reporting_date__range=[week_start, week_end])
        else:
            # Default to current week
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            queryset = queryset.filter(reporting_date__range=[week_start, week_end])
        
        # Apply district filter
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        
        # Apply PO filter
        if po_id:
            queryset = queryset.filter(agent_id=po_id)
        
        # Calculate summary statistics
        summary = {
            'total_reports': queryset.count(),
            'total_area': queryset.aggregate(total=Sum('area_covered_ha'))['total'] or 0,
            'total_ras': queryset.aggregate(total=Sum('no_rehab_assistants'))['total'] or 0,
            'avg_area_per_report': queryset.aggregate(avg=Avg('area_covered_ha'))['avg'] or 0,
            'submitted_reports': queryset.filter(status=1).count(),
            'approved_reports': queryset.filter(status=2).count(),
            'pending_reports': queryset.filter(status=0).count(),
            'rejected_reports': queryset.filter(status=3).count(),
            'completion_rate': (queryset.filter(status=2).count() / queryset.count() * 100) if queryset.count() > 0 else 0,
        }
        
        # Get top activities this week
        top_activities = queryset.values(
            'main_activity__main_activity'
        ).annotate(
            total_area=Sum('area_covered_ha'),
            report_count=Count('id')
        ).order_by('-total_area')[:5]
        
        summary['top_activities'] = list(top_activities)
        
        # Get district performance
        district_performance = queryset.values(
            'district__name'
        ).annotate(
            total_area=Sum('area_covered_ha'),
            report_count=Count('id'),
            total_ras=Sum('no_rehab_assistants')
        ).order_by('-total_area')[:5]
        
        summary['district_performance'] = list(district_performance)
        
        return JsonResponse({'success': True, 'data': summary})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_trends(request):
    """Get weekly trends for charts"""
    try:
        # Get last 12 weeks of data
        today = timezone.now().date()
        trends = []
        
        for i in range(11, -1, -1):
            week_start = today - timedelta(days=today.weekday() + 7 * i)
            week_end = week_start + timedelta(days=6)
            week_num = week_start.isocalendar()[1]
            year = week_start.year
            
            week_data = DailyReportingModel.objects.filter(
                delete_field='no',
                reporting_date__range=[week_start, week_end]
            ).aggregate(
                total_area=Sum('area_covered_ha'),
                report_count=Count('id'),
                total_ras=Sum('no_rehab_assistants')
            )
            
            trends.append({
                'week': f"W{week_num}",
                'year': year,
                'start_date': week_start.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'total_area': float(week_data['total_area'] or 0),
                'report_count': week_data['report_count'] or 0,
                'total_ras': week_data['total_ras'] or 0
            })
        
        return JsonResponse({'success': True, 'data': trends})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_po_performance(request):
    """Get PO performance metrics"""
    try:
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        
        if not week_start or not week_end:
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
        
        # Get PO performance
        po_performance = DailyReportingModel.objects.filter(
            delete_field='no',
            reporting_date__range=[week_start, week_end]
        ).values(
            'agent_id',
            'agent__first_name',
            'agent__last_name',
            'agent__staffid',
            'district__name'
        ).annotate(
            total_reports=Count('id'),
            total_area=Sum('area_covered_ha'),
            total_ras=Sum('no_rehab_assistants'),
            approved_reports=Count('id', filter=Q(status=2)),
            pending_reports=Count('id', filter=Q(status__in=[0, 1])),
            avg_area_per_report=Avg('area_covered_ha')
        ).order_by('-total_area')[:10]
        
        data = []
        for po in po_performance:
            data.append({
                'po_id': po['agent_id'],
                'po_name': f"{po['agent__first_name'] or ''} {po['agent__last_name'] or ''}".strip() or 'Unknown',
                'staff_id': po['agent__staffid'] or 'N/A',
                'district': po['district__name'] or 'N/A',
                'total_reports': po['total_reports'],
                'total_area': float(po['total_area'] or 0),
                'total_ras': po['total_ras'] or 0,
                'approved_reports': po['approved_reports'],
                'pending_reports': po['pending_reports'],
                'approval_rate': (po['approved_reports'] / po['total_reports'] * 100) if po['total_reports'] > 0 else 0,
                'avg_area': float(po['avg_area_per_report'] or 0)
            })
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_activity_breakdown(request):
    """Get activity breakdown for current week"""
    try:
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        
        if not week_start or not week_end:
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
        
        # Get activity breakdown
        activity_breakdown = DailyReportingModel.objects.filter(
            delete_field='no',
            reporting_date__range=[week_start, week_end]
        ).values(
            'main_activity__main_activity',
            'activity__sub_activity'
        ).annotate(
            total_area=Sum('area_covered_ha'),
            report_count=Count('id'),
            total_ras=Sum('no_rehab_assistants')
        ).order_by('-total_area')
        
        # Group by main activity
        main_activities = {}
        for item in activity_breakdown:
            main_act = item['main_activity__main_activity'] or 'Other'
            if main_act not in main_activities:
                main_activities[main_act] = {
                    'name': main_act,
                    'total_area': 0,
                    'report_count': 0,
                    'total_ras': 0,
                    'sub_activities': []
                }
            
            main_activities[main_act]['total_area'] += float(item['total_area'] or 0)
            main_activities[main_act]['report_count'] += item['report_count'] or 0
            main_activities[main_act]['total_ras'] += item['total_ras'] or 0
            
            main_activities[main_act]['sub_activities'].append({
                'name': item['activity__sub_activity'] or 'Unknown',
                'area': float(item['total_area'] or 0),
                'reports': item['report_count'] or 0,
                'ras': item['total_ras'] or 0
            })
        
        return JsonResponse({
            'success': True, 
            'data': list(main_activities.values())
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_district_summary(request):
    """Get district-wise summary"""
    try:
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        
        if not week_start or not week_end:
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
        
        # Get district summary
        district_summary = DailyReportingModel.objects.filter(
            delete_field='no',
            reporting_date__range=[week_start, week_end],
            district__isnull=False
        ).values(
            'district__id',
            'district__name',
            'district__region__region'
        ).annotate(
            total_reports=Count('id'),
            total_area=Sum('area_covered_ha'),
            total_ras=Sum('no_rehab_assistants'),
            unique_pos=Count('agent_id', distinct=True),
            approved_reports=Count('id', filter=Q(status=2))
        ).order_by('-total_area')
        
        data = []
        for district in district_summary:
            data.append({
                'id': district['district__id'],
                'name': district['district__name'] or 'Unknown',
                'region': district['district__region__region'] or 'Unknown',
                'total_reports': district['total_reports'],
                'total_area': float(district['total_area'] or 0),
                'total_ras': district['total_ras'] or 0,
                'unique_pos': district['unique_pos'] or 0,
                'approved_reports': district['approved_reports'],
                'approval_rate': (district['approved_reports'] / district['total_reports'] * 100) if district['total_reports'] > 0 else 0
            })
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_reports(request):
    """Get detailed weekly reports for DataTable"""
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        district_id = request.GET.get('district_id')
        po_id = request.GET.get('po_id')
        activity_id = request.GET.get('activity_id')
        status = request.GET.get('status')
        
        # Base queryset
        queryset = DailyReportingModel.objects.select_related(
            'agent', 'main_activity', 'activity', 'farm', 
            'community', 'district'
        ).filter(
            delete_field='no'
        )
        
        # Apply date filter
        if week_start and week_end:
            queryset = queryset.filter(reporting_date__range=[week_start, week_end])
        else:
            # Default to current week
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            queryset = queryset.filter(reporting_date__range=[week_start, week_end])
        
        # Apply other filters
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        if po_id:
            queryset = queryset.filter(agent_id=po_id)
        if activity_id:
            queryset = queryset.filter(main_activity_id=activity_id)
        if status and status.strip():
            queryset = queryset.filter(status=status)
        
        # Apply search
        if search_value:
            queryset = queryset.filter(
                Q(agent__first_name__icontains=search_value) |
                Q(agent__last_name__icontains=search_value) |
                Q(farm_ref_number__icontains=search_value) |
                Q(main_activity__main_activity__icontains=search_value) |
                Q(district__name__icontains=search_value)
            )
        
        # Get total count
        total_records = queryset.count()
        
        # Apply pagination
        reports = queryset.order_by('-reporting_date')[start:start + length]
        
        # Prepare data
        data = []
        for report in reports:
            data.append({
                'id': report.id,
                'reporting_date': report.reporting_date.strftime('%Y-%m-%d') if report.reporting_date else '',
                'agent_name': f"{report.agent.first_name} {report.agent.last_name}" if report.agent else 'N/A',
                'main_activity': report.main_activity.main_activity if report.main_activity else 'N/A',
                'sub_activity': report.activity.sub_activity if report.activity else 'N/A',
                'farm_ref_number': report.farm_ref_number or 'N/A',
                'area_covered_ha': float(report.area_covered_ha) if report.area_covered_ha else 0,
                'no_rehab_assistants': report.no_rehab_assistants or 0,
                'status': report.status,
                'status_display': get_status_display(report.status),
                'district': report.district.name if report.district else 'N/A',
                'community': report.community.name if report.community else 'N/A',
                'remark': report.remark or '',
            })
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data,
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_export(request):
    """Export weekly monitoring data to CSV/Excel"""
    try:
        import csv
        from django.http import HttpResponse
        
        # Get filter parameters
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        format_type = request.GET.get('format', 'csv')
        
        if not week_start or not week_end:
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
        
        # Get data
        reports = DailyReportingModel.objects.select_related(
            'agent', 'main_activity', 'activity', 'district', 'community'
        ).filter(
            delete_field='no',
            reporting_date__range=[week_start, week_end]
        ).order_by('-reporting_date')
        
        # Create response
        filename = f"weekly_monitoring_{week_start}_to_{week_end}.{format_type}"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Project Officer', 'District', 'Community', 
            'Main Activity', 'Sub Activity', 'Farm Reference',
            'Area (ha)', '# RAs', 'Status', 'Remarks'
        ])
        
        for report in reports:
            writer.writerow([
                report.reporting_date,
                f"{report.agent.first_name} {report.agent.last_name}" if report.agent else 'N/A',
                report.district.name if report.district else 'N/A',
                report.community.name if report.community else 'N/A',
                report.main_activity.main_activity if report.main_activity else 'N/A',
                report.activity.sub_activity if report.activity else 'N/A',
                report.farm_ref_number or 'N/A',
                f"{report.area_covered_ha:.2f}" if report.area_covered_ha else '0.00',
                report.no_rehab_assistants or 0,
                get_status_display(report.status),
                report.remark or ''
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

def get_status_display(status):
    status_map = {
        0: 'Pending',
        1: 'Submitted',
        2: 'Approved',
        3: 'Rejected'
    }
    return status_map.get(status, 'Unknown')


@csrf_exempt
@require_http_methods(["GET"])
def weekly_monitoring_export(request):
    """Export weekly monitoring data to CSV/Excel"""
    try:
        format_type = request.GET.get('format', 'csv')
        week_start = request.GET.get('week_start')
        week_end = request.GET.get('week_end')
        district_id = request.GET.get('district_id')
        po_id = request.GET.get('po_id')
        activity_id = request.GET.get('activity_id')
        
        if not week_start or not week_end:
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
        
        # Base queryset
        queryset = DailyReportingModel.objects.select_related(
            'agent', 'main_activity', 'activity', 'district', 'community'
        ).filter(
            delete_field='no',
            reporting_date__range=[week_start, week_end]
        )
        
        # Apply filters
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        if po_id:
            queryset = queryset.filter(agent_id=po_id)
        if activity_id:
            queryset = queryset.filter(main_activity_id=activity_id)
        
        if format_type == 'csv':
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="weekly_monitoring_{week_start}_to_{week_end}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Date', 'Project Officer', 'Staff ID', 'District', 'Community',
                'Main Activity', 'Sub Activity', 'Farm Reference',
                'Area (ha)', '# RAs', 'Group Work', 'People in Group',
                'Status', 'Remarks'
            ])
            
            for report in queryset:
                writer.writerow([
                    report.reporting_date,
                    f"{report.agent.first_name} {report.agent.last_name}" if report.agent else 'N/A',
                    report.agent.staffid if report.agent else 'N/A',
                    report.district.name if report.district else 'N/A',
                    report.community.name if report.community else 'N/A',
                    report.main_activity.main_activity if report.main_activity else 'N/A',
                    report.activity.sub_activity if report.activity else 'N/A',
                    report.farm_ref_number or 'N/A',
                    f"{report.area_covered_ha:.2f}" if report.area_covered_ha else '0.00',
                    report.no_rehab_assistants or 0,
                    report.group_work or 'No',
                    report.number_of_people_in_group or 0,
                    get_status_display(report.status),
                    report.remark or ''
                ])
            
            return response
            
        elif format_type == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from django.http import HttpResponse
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Weekly Monitoring"
                
                # Headers
                headers = [
                    'Date', 'Project Officer', 'Staff ID', 'District', 'Community',
                    'Main Activity', 'Sub Activity', 'Farm Reference',
                    'Area (ha)', '# RAs', 'Group Work', 'People in Group',
                    'Status', 'Remarks'
                ]
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3b5e7e", end_color="3b5e7e", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data rows
                for row, report in enumerate(queryset, 2):
                    ws.cell(row=row, column=1, value=report.reporting_date.strftime('%Y-%m-%d') if report.reporting_date else '')
                    ws.cell(row=row, column=2, value=f"{report.agent.first_name} {report.agent.last_name}" if report.agent else 'N/A')
                    ws.cell(row=row, column=3, value=report.agent.staffid if report.agent else 'N/A')
                    ws.cell(row=row, column=4, value=report.district.name if report.district else 'N/A')
                    ws.cell(row=row, column=5, value=report.community.name if report.community else 'N/A')
                    ws.cell(row=row, column=6, value=report.main_activity.main_activity if report.main_activity else 'N/A')
                    ws.cell(row=row, column=7, value=report.activity.sub_activity if report.activity else 'N/A')
                    ws.cell(row=row, column=8, value=report.farm_ref_number or 'N/A')
                    ws.cell(row=row, column=9, value=float(report.area_covered_ha) if report.area_covered_ha else 0)
                    ws.cell(row=row, column=10, value=report.no_rehab_assistants or 0)
                    ws.cell(row=row, column=11, value=report.group_work or 'No')
                    ws.cell(row=row, column=12, value=report.number_of_people_in_group or 0)
                    ws.cell(row=row, column=13, value=get_status_display(report.status))
                    ws.cell(row=row, column=14, value=report.remark or '')
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="weekly_monitoring_{week_start}_to_{week_end}.xlsx"'
                
                wb.save(response)
                return response
                
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'openpyxl is not installed. Please install it for Excel export.'
                }, status=500)
        
        return JsonResponse({'success': False, 'message': 'Invalid format'}, status=400)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)